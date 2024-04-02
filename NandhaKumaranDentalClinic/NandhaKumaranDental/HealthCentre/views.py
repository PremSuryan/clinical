from django.contrib.auth import authenticate, login, logout
from django.http import HttpResponse, HttpResponseRedirect, JsonResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from .models import Doctor, Patient, Prescription, passwordHasher, emailHasher, Appointment, Medicine
from django.db.models import Count, Q
from django.contrib.auth.decorators import login_required
from .forms import AppointmentSet, AppointmentSetForm, AppointmentForm
from datetime import datetime, time, timedelta
import time
from django.utils import timezone
from django.shortcuts import render
import threading
import sys, os
import pyautogui
import openpyxl
from django.db import connections
from django.apps import apps
import requests
import zipfile

if 'runserver' in sys.argv:
    from .Whatsapptestfile import whatsappApi, openWhatsapp, whatsappApiEdit, whatsappMedia

def updateExcel():
    while True:
        xlPath = os.curdir #"D:\Dental-Software-Backup\Dental-Software"
        allfilesinpath = os.listdir(xlPath)
        xlFile = [file for file in allfilesinpath if file.lower().startswith('databasetables.xlsx')]
        if not xlFile:
            workbook = openpyxl.Workbook()
            for sheetIndex, model in enumerate(apps.get_models()):
                if model.__name__ in [ 'Patient', 'Prescription', 'Appointment']:

                    if model.__name__ == 'Prescription':
                        excludedColumns = ['timestamp', 'doctor', 'patient']
                        worksheet = workbook.create_sheet(title='sheet2')
                        
                    elif model.__name__ == 'Appointment':
                        excludedColumns = ['time', 'date', 'AppointmentTimeStamp']
                        worksheet = workbook.create_sheet(title='sheet3')
                        
                    else:
                        excludedColumns = []
                        worksheet = workbook.create_sheet(title='sheet1')

                    allColumns = [field.name for field in model._meta.fields]
                    includedColumns = [col for col in allColumns if col not in excludedColumns]
                    
                    with connections['default'].cursor() as cursor:
                        cursor.execute(f'SELECT {",".join(includedColumns)} FROM {model._meta.db_table}')
                        rows = cursor.fetchall()
                        
                    headers = includedColumns
                    worksheet.append(headers)

                    for row in rows:
                        worksheet.append(row)
            try : 
                workbook.save('databasetables.xlsx')
                workbook.close()
            except PermissionError:
                time.sleep(1)
        else:
            try:
                workbookExisting = openpyxl.load_workbook('databasetables.xlsx')
            except (zipfile.BadZipfile) as exception:
                time.sleep(1)
                os.remove('databasetables.xlsx')
                updateExcel()
           
            for sheetIndex, model in enumerate(apps.get_models()):
                if model.__name__ in [ 'Patient', 'Prescription', 'Appointment']:
                    
                    if model.__name__ == 'Prescription':
                        excludedColumns = ['timestamp', 'doctor', 'patient']
                        worksheetExisting = workbookExisting.get_sheet_by_name('sheet2')
                        
                    elif model.__name__ == 'Appointment':
                        excludedColumns = ['time', 'date', 'AppointmentTimeStamp']
                        worksheetExisting = workbookExisting.get_sheet_by_name('sheet3')
                        
                    else:
                        excludedColumns = []
                        worksheetExisting = workbookExisting.get_sheet_by_name('sheet1')
                        
                    worksheetExisting.delete_rows(1, worksheetExisting.max_row)
                    allColumns = [field.name for field in model._meta.fields]
                    includedColumns = [col for col in allColumns if col not in excludedColumns]
                    
                    with connections['default'].cursor() as cursor:
                        cursor.execute(f'SELECT {",".join(includedColumns)} FROM {model._meta.db_table}')
                        rows = cursor.fetchall()
                    headers = includedColumns
                    worksheetExisting.append(headers)

                    for row in rows:
                        worksheetExisting.append(row)
            try:
                workbookExisting.save('databasetables.xlsx')
                workbookExisting.close()
            except PermissionError:
                time.sleep(1)

def index(request):
    """ Function for displaying main page of website. """
    
    # Editing response headers so as to ignore cached versions of pages
    response = render(request,"HealthCentre/index.html")
    return responseHeadersModifier(response)

def register(request):
    """ Function for registering a student into the portal. """

    # If the user wants the page to get displayed

    if request.method == "GET":
        # Editing response headers so as to ignore cached versions of pages

        response =  render(request,"HealthCentre/registrationPortal.html")

        return responseHeadersModifier(response)
    
    # If the user wants to submit his/her information

    elif request.method == "POST":
        # Extracting the user information from the post request
        userFirstName = request.POST["userFirstNam"]
        userLastName = request.POST["userLastName"]
        userEmail = request.POST["userEmail"]
        userRollNo = request.POST["userRollNo"]
        userAddress = request.POST["userAddress"]
        userContactNo = request.POST["userContactNo"]
        userPassword = request.POST["userPassword"]
        
        userType = request.POST['userType']
        if userType == 'patient':
            userConfirmPassword = userPassword
        elif userType == 'doctor':
            userConfirmPassword = request.POST["userConfirmPassword"]
        # If both the passwords match
        if userPassword == userConfirmPassword:

            name = userFirstName + " " + userLastName

            

            # handleSubmit(request)
            # def handleSubmit(request):
            
            
            # Creating a patient object and saving insdie the database if patient is selected 
            
            if userType == 'patient':
                # patient = Patient(rollNumber=request.POST['rollNumber'])
                # Encrypting password to store inside database
                passwordHash = userPassword

                # Encrypting email to store inside database
                emailHash = emailHasher(userEmail)
                patient = Patient(name = name,rollNumber = userRollNo, email = userEmail, passwordHash = passwordHash, address = userAddress, contactNumber = userContactNo, emailHash = emailHash )
                patient.save()
                    
            # Creating a patient object and saving insdie the database if patient is selected
            elif userType == 'doctor':
                doctor = Doctor(name = name, specialization= userRollNo, email = userEmail, passwordHash = passwordHash, address = userAddress, contactNumber = userContactNo, emailHash = emailHash)
                doctor.save()


            
            

            # Creating a patient object and saving insdie the database
            # patient = Patient(name = name,rollNumber = userRollNo, email = userEmail, passwordHash = passwordHash, address = userAddress, contactNumber = userContactNo, emailHash = emailHash )
            # patient.save()

            # Storing success message in the context variable
            context = {
                "userType" : userType,
                "message":"User Registration Successful. Please Login."
            }

            # Editing response headers so as to ignore cached versions of pages
            response = render(request, "HealthCentre/registrationPortal.html",context)
            return responseHeadersModifier(response)

        # If the passwords given don't match
        else:
            # Storing failure message in the context variable
            context = {
                "message":"Passwords do not match.Please register again."
            }

            # Editing response headers so as to ignore cached versions of pages
            response = render(request,"HealthCentre/registrationPortal.html",context)
            return responseHeadersModifier(response)

    # For any other method of request, sending back the registration page.
    else:

        # Editing response headers so as to ignore cached versions of pages
        response = render(request,"HealthCentre/registrationPortal.html")
        return responseHeadersModifier(response)


def doctors(request):
    """Function to send information about the doctors available to the user. """

    # Storing doctors available in the context variable
    context = {
        "doctors" : Doctor.objects.all()
    }

    # Editing response headers so as to ignore cached versions of pages
    response = render(request,"HealthCentre/doctors.html",context)
    return responseHeadersModifier(response)


def login(request):
    """ Function for logging in the user. """

    # Calling session variables checker
    request = requestSessionInitializedChecker(request)

    # If the request method is post
    if request.method == "GET":
        try:

            # If the user is already logged in inside of his sessions, and is a doctor, then no authentication required
            if request.session['isLoggedIn'] and request.session['isDoctor']:
                # Accessing the doctor user and all his/her records
                doctor = Doctor.objects.get(emailHash = request.session['userEmail'])
                records = doctor.doctorRecords.all()
                # Getting the count of the new prescriptions pending
                numberNewPendingPrescriptions = doctor.doctorRecords.aggregate(newnewPendingPrescriptions = Count('pk', filter =( Q(isNew = True) & Q(isCompleted = False) ) ))['newnewPendingPrescriptions']
                # Storing the same inside the session variables
                request.session['numberNewPrescriptions'] = numberNewPendingPrescriptions
                # Storing the required information inside the context variable
                context = {
                    "message" : "Successfully Logged In.",
                    "isAuthenticated" : True,
                    "user": records.order_by('-timestamp'),
                    "prescriptions" : Prescription.objects.all().order_by('timestamp'),
                    "prescMedicine" : Medicine.objects.all().order_by('id')
                }
                
                # Editing response headers so as to ignore cached versions of pages
                response = render(request,"HealthCentre/prescriptionPortal.html", context)
                return responseHeadersModifier(response)
            
            # If the user is already logged in inside of his sessions, and is a patient, then no authentication required
            elif request.session['isLoggedIn'] and (not request.session['isDoctor']):

                # Accessing the patient user and all his/her records
                patient = Patient.objects.get(emailHash = request.session['userEmail'])
                records = patient.patientRecords.all()

                # Getting the count of the new prescriptions pending
                numberNewPrescriptions = patient.patientRecords.aggregate(newCompletedPrescriptions = Count('pk', filter =( Q(isNew = True) & Q(isCompleted = True) ) ) )['newCompletedPrescriptions']

                # Storing the same inside the session variables
                request.session['numberNewPrescriptions'] = numberNewPrescriptions

                # Updating the completed records
                for record in records:
                    if record.isCompleted:
                        record.isNew = False
                        record.save()

                # Storing the required information inside the context variable
                context = {
                    "message" : "Successfully Logged In.",
                    "isAuthenticated" : True,
                    "user": records.order_by('-timestamp')
                    }

                # Editing response headers so as to ignore cached versions of pages
                # response = render(request,"HealthCentre/prescriptionPortal.html")
                response = render(request,"HealthCentre/userPatientProfilePortal.html", context)
                return responseHeadersModifier(response)

            else:
                # Editing response headers so as to ignore cached versions of pages
                response = render(request,"HealthCentre/loginPortal.html")
                return responseHeadersModifier(response)

        # If any error occurs, sending back a new blank page
        except:

            # Editing response headers so as to ignore cached versions of pages
            response = render(request,"HealthCentre/loginPortal.html")
            return responseHeadersModifier(response)

    # If the request method is post
    elif request.method == "POST":

        # Extracting the user information from the post request
        userName = request.POST["useremail"]
        userPassword = request.POST["userpassword"]

        # If such a patient exists
        try:
            patient = Patient.objects.get(email = userName)

            # Storing required session information
            request.session['isDoctor'] = False

        # Otherwise trying if a doctor exists
        except Patient.DoesNotExist:
            try:
                doctor = Doctor.objects.get(email = userName)

                # Storing required session information
                request.session['isDoctor'] = True     

            # If no such doctor or patient exists
            except Doctor.DoesNotExist:

                # Storing message inside context variable
                context = {
                    "message":"User does not exist.Please register first."
                }

                # Editing response headers so as to ignore cached versions of pages
                response = render(request,"HealthCentre/loginPortal.html", context)
                return responseHeadersModifier(response)

        # Getting the hash of user inputted password
        passwordHash = passwordHasher(userPassword)

        # If the logged in user is a doctor
        if request.session['isDoctor']:
            
            # Accessing all records of doctor
            records = doctor.doctorRecords.all()

            # Getting the count of new prescriptions
            numberNewPendingPrescriptions = doctor.doctorRecords.aggregate(newPendingPrescriptions = Count('pk', filter =( Q(isNew = True) & Q(isCompleted = False) ) ))['newPendingPrescriptions']

            # Storing the same inside request variable
            request.session['numberNewPrescriptions'] = numberNewPendingPrescriptions

            # If the inputted hash and the original user password hash match
            if passwordHash == doctor.passwordHash:

                # Storing required information in session variable of request
                request.session['isLoggedIn'] = True
                request.session['userEmail'] = doctor.emailHash
                request.session['Name'] = doctor.name

                # Redirecting to avoid form resubmission
                # Redirecting to home page
                # Editing response headers so as to ignore cached versions of pages
                # response = render(request,"HealthCentre/userDoctorProfilePortal.html")
                # response = HttpResponseRedirect(reverse('onlineprescription'))
                # response = render(request,"HealthCentre/prescriptionPortal.html")
                response = HttpResponseRedirect(reverse('index'))
                return responseHeadersModifier(response)

            # Else if the password inputted is worng and doesn't match
            else:

                # Storing message inside context variable
                context = {
                    "message":"Invalid Credentials.Please Try Again."
                }

                # Editing response headers so as to ignore cached versions of pages
                response = render(request,"HealthCentre/loginPortal.html", context)
                return responseHeadersModifier(response)

        # Otherwise if the user is a patient
        else:

            # Accessing all records of patient
            records = patient.patientRecords.all()

            # Getting the count of new prescriptions
            numberNewPrescriptions = patient.patientRecords.aggregate(newCompletedPrescriptions = Count('pk', filter =( Q(isNew = True) & Q(isCompleted = True) ) ))['newCompletedPrescriptions']

            # Storing the same inside request variable
            request.session['numberNewPrescriptions'] = numberNewPrescriptions

            # Updating the completed records
            for record in records:
                if record.isCompleted :
                    record.isNew = False
                    record.save()

            # If the inputted hash and the original user password hash match
            if passwordHash == patient.passwordHash:

                # Storing required information in session variable of request
                request.session['isLoggedIn'] = True
                request.session['userEmail'] = patient.emailHash
                request.session['Name'] = patient.name
                request.session['isDoctor'] = False

                # Redirecting to avoid form resubmission
                # Redirecting to home page
                # Editing response headers so as to ignore cached versions of pages
                # response = render(request,"HealthCentre/userPatientProfilePortal.html")
                # response = render(request, "HealthCentre/prescriptionportal.html")
                # response = HttpResponseRedirect(reverse('onlineprescription'))
                response = HttpResponseRedirect(reverse('index'))
                return responseHeadersModifier(response)

            # Else if the password inputted is worng and doesn't match
            else:

                # Storing message inside context variable
                context = {
                    "message":"Invalid Credentials.Please Try Again."
                }

                # Editing response headers so as to ignore cached versions of pages
                response = render(request,"HealthCentre/loginPortal.html", context)
                return responseHeadersModifier(response)
    # For any other method of access, returning a new blank login page
    else:
        response = render(request,"HealthCentre/loginPortal.html")
        return responseHeadersModifier(response)

def emergency(request):
    """ Funtion for emergency situations, for requesting an ambulance."""

    # If the request method is get
    if request.method == "GET":

        # Editing response headers so as to ignore cached versions of pages
        response = render(request,"HealthCentre/emergencyPortal.html")
        return responseHeadersModifier(response)

    # If the request method is post and the user is submitting information
    elif request.method == "POST":

        # Extracting the emergency location from the post request
        emergencyLocation = request.POST['emergencyLocation']

        # Giving emergency message to server, can also be connected to IOT devices for alarms
        # If the emergency location text is not an empty string
        if emergencyLocation != "":

            # Printing information and notifying inside of server terminal
            print("------------------------------------------------------------------------")
            print("\n\nEMERGENCY !! AMBULANCE REQUIRED AT " + emergencyLocation + " !!\n\n")
            print("------------------------------------------------------------------------")

            # Storing information inside of context variable
            context = {
                "message" : "Ambulance reaching " + emergencyLocation + " in 2 minutes."
            }

            # Editing response headers so as to ignore cached versions of pages
            response = render(request, "HealthCentre/emergencyPortal.html", context)
            return responseHeadersModifier(response)

        # Else if the emergency location is an empty string
        else:

            # Storing message inside context variable
            context = {
                "message" : "No location entered.Invalid input."
            }

            # Editing response headers so as to ignore cached versions of pages
            response = render(request, "HealthCentre/emergencyPortal.html", context)
            return responseHeadersModifier(response)

    # For any other method of access, returning a new blank emergency page
    else:

        # Editing response headers so as to ignore cached versions of pages
        response = render(request,"HealthCentre/emergencyPortal.html")
        return responseHeadersModifier(response)

def logout(request):
    """Function to log out the user."""
    # Erasing all the information of the session variables if user is logged out
    request.session['isDoctor'] = ""
    request.session['isLoggedIn'] = False
    request.session['userEmail'] = ""
    request.session['Name'] = ""
    request.session['numberNewPrescriptions'] = ""
    request.session['writeNewPrescription'] = False
    request.session['createNewAppointment'] = False

    # Redirecting to avoid form resubmission
    # Redirecting to home page
    # Editing response headers so as to ignore cached versions of pages
    response = HttpResponseRedirect(reverse('login'))
    return responseHeadersModifier(response)

def contactus(request):
    """Function to display contact information."""

    # Editing response headers so as to ignore cached versions of pages
    response = render(request, "HealthCentre/contactus.html")
    return responseHeadersModifier(response)


def doctorappointmentsfalse(request):
    if request.method == 'GET':
        # request.session['goToAppointmentsPage'] = True
        request.session['createNewAppointment'] = True
        if request.session['isLoggedIn'] and request.session['isDoctor'] and request.session['createNewAppointment']:
            # Accessing the doctor user and all his/her records
            # request.session['CreatenewAppointment'] = False
            doctor = Doctor.objects.get(emailHash = request.session['userEmail'])
            records = doctor.doctorRecords.all()
            # Getting the count of the new prescriptions pending
            numberNewPendingPrescriptions = doctor.doctorRecords.aggregate(newnewPendingPrescriptions = Count('pk', filter =( Q(isNew = True) & Q(isCompleted = False) ) ))['newnewPendingPrescriptions']

            # Storing the same inside the session variables
            request.session['numberNewPrescriptions'] = numberNewPendingPrescriptions

            # Storing the required information inside the context variable
            context = {
                "message" : "Successfully Logged In.",
                "isAuthenticated" : True,
                "user": records.order_by('-timestamp'),
                "Appointments" : Appointment.objects.all().order_by('-date')
            }
            response = render(request,"HealthCentre/appointmentsPortal.html", context)
            return responseHeadersModifier(response)
        
def doctorappointments(request):
    if request.method == 'GET':
        request.session['goToAppointmentsPage'] = True
        request.session['appointmentEdit'] = False
        # request.session['createNewAppointment'] = True
        form = AppointmentForm()
        model = AppointmentForm()
        # form = AppointmentForm(request.POST or None)
        hour = range(00, 24)
        minute = range(00, 60)
        date = range(1, 32)
        month = range(1, 13)
        year = range(int(datetime.now().year), 2099)
        
        context = {'form': form, 
                    'model': model,
                    'hours': hour,
                    'dates' : date,
                    'months' : month,
                    'years' : year,
                    'minutes': minute,
                    "patients" : Patient.objects.all().order_by('id'),
                    # "prescPatients" : Appointment.objects.all().order_by('id')
                    }
        response = render(request, 'HealthCentre/NewAppointment.html', context)
        return responseHeadersModifier(response)
    if request.method == 'POST':
        if request.session['goToAppointmentsPage']:
            if request.POST['selectedPatient'] == "":
                appointmentPatient = request.POST['PatientNameForAppointment']
                # patient = Patient.objects.create(name=prescpatient)
            else:
                appointmentPatient = request.POST['selectedPatient']
                # prescpatient = request.POST['selectedPatient']
                patient_id = request.POST['selectedPatient'] 
                patient = Patient.objects.get(name=patient_id)
            appointmentTime = request.POST['EnterTimeHour'].zfill(2) + request.POST['EnterTimeMinute'].zfill(2)
            datetimeObject = datetime.strptime(appointmentTime, "%H%M")
            appointmentDate = request.POST['EnterDate'] + request.POST['EnterDateMonth'] + request.POST['EnterYear']
            dateobject = datetime.strptime(appointmentDate, "%d%m%Y")
            appointmentNotes = request.POST['AppointmentDescription']
            appointmentDoctor = request.session['Name']
            appointmentSubject = "subject"
            doctor_id = request.session['Name']
            doctorid = Doctor.objects.get(name=doctor_id)
            appointment = Appointment(time = datetimeObject, date = dateobject, subject = appointmentSubject, notes = appointmentNotes,
                                        appointmentpatient = appointmentPatient, appointmentdoctor = appointmentDoctor, doctorPres = doctorid,
                                        patientPres = patient)
            appointment.save()
            doctor = Doctor.objects.get(emailHash = request.session['userEmail'])
            records = doctor.doctorRecords.all()
            context = {
            "message" : "Successfully Logged In.",
            "isAuthenticated" : True,
            "user": records.order_by('-timestamp'),
            "Appointments" : Appointment.objects.all().order_by('-date')
        }
        # whatsappNotification()
        response = HttpResponseRedirect(reverse('doctorappointmentsfalse'))
        return responseHeadersModifier(response)

def editAppointments(request, pk):
    request.session['appointmentEdit'] = True
    # record = get_object_or_404(Appointment, pk = record_id)
    appointment = Appointment.objects.get(id=pk)
    form = AppointmentForm(instance=appointment)
    if request.method == "POST":
        # form = AppointmentForm(request.POST, instance=appointment)
        # if form.is_valid():
            # form.save()
        doctor = Doctor.objects.get(emailHash = request.session['userEmail'])
        records = doctor.doctorRecords.all()
        appointObject = Appointment.objects.get(id=pk)
        if request.POST['selectedPatient'] == "":
            appointObject.appointmentpatient = request.POST['PatientNameForAppointment']
            # patient = Patient.objects.create(name=prescpatient)
        else:
            appointObject.appointmentpatient = request.POST['selectedPatient']
            # prescpatient = request.POST['selectedPatient']
            # patient_id = request.POST['selectedPatient'] 
            # patient = Patient.objects.get(name=patient_id)
        
        appointmentTime = request.POST['EnterTimeHour'].zfill(2) + request.POST['EnterTimeMinute'].zfill(2)
        appointmentTime = datetime.strptime(appointmentTime, "%H%M")
        appointObject.time = appointmentTime
        appointmentDate = request.POST['EnterDate'].zfill(2) + request.POST['EnterDateMonth'].zfill(2) + request.POST['EnterYear'].zfill(2)
        appointmentDate = datetime.strptime(appointmentDate, "%d%m%Y")
        appointObject.date = appointmentDate
        appointObject.notes = request.POST['AppointmentDescription']
        appointObject.appointmentdoctor = request.session['Name']
        appointObject.subject = "subject"
        appointObject.save()
        patientDetails = Patient.objects.get(name=appointObject.appointmentpatient)
        patientNumber = patientDetails.contactNumber 
        whatsappApiEdit(appointObject.appointmentpatient, patientNumber, appointObject.time, appointObject.date)
        # appointment = Appointment(time = datetimeObject, date = dateobject, subject = appointmentSubject, notes = appointmentNotes,
        #                             appointmentpatient = appointmentPatient, appointmentdoctor = appointmentDoctor)
        # appointment.save()
        doctor = Doctor.objects.get(emailHash = request.session['userEmail'])
        records = doctor.doctorRecords.all()
        context = {
        "message" : "Successfully Logged In.",
        "isAuthenticated" : True,
        "user": records.order_by('-timestamp'),
        "Appointments" : Appointment.objects.all().order_by('-date')
    }
        # response = HttpResponseRedirect(request, 'HealthCentre/appointmentsPortal.html', context)
        response = HttpResponseRedirect(reverse('doctorappointmentsfalse'))
        return responseHeadersModifier(response)
        
    # hour = Appointment.objects.get()
    # hour = Appointment.objects.get(time=time)
    hour = range(00, 24)
    minute = range(00, 60)
    date = range(1, 32)
    month = range(1, 13)
    year = range(int(datetime.now().year), 2099)
    appointmentObject = Appointment.objects.get(id=pk)
    patient = appointmentObject.appointmentpatient
    datobject = appointmentObject.date
    dateobject = datetime.strftime(datobject, "%d")
    # monthobject = appointmentObject.date
    monthobject = datetime.strftime(datobject, "%m")
    # yearobject = appointmentObject.date
    yearobject = datetime.strftime(datobject, "%Y")
    timeobject = datetime.combine(appointmentObject.date, appointmentObject.time) 
    minuteobject = datetime.strftime(timeobject, "%M")
    hourobject = datetime.strftime(timeobject, "%H")
    appointmentnotes = appointmentObject.notes
    context = {
        
            'hours': hour,
            'editDate' : dateobject,
            'dates' : date,
            'editMonth' : monthobject,
            'months' : month,
            'editYear' : yearobject,
            'years' : year,
            'editMinute' : minuteobject,
            'minutes': minute,
            'editHour' : hourobject,
            "patients" : patient,
            "editNotes" : appointmentnotes,
            "pats" : Patient.objects.all().order_by('id'),
            "prescPatients" : Appointment.objects.all().order_by('id'),
            'form' : form
    }
    # appointmentObject.date = datetime.strptime((request.POST.get('EnterDate') + request.POST.get('EnterDateMonth') + request.POST.get('EnterYear')), "%m%d%Y") 
    # appointmentObject.date = datetime.strptime(appointmentObject.date, "%m%d%Y")
    # appointmentObject.save()
    # return render(request,'HealthCentre/NewAppointment.html', context)
    response = render(request,'HealthCentre/NewAppointment.html', context)
    return responseHeadersModifier(response)

def deleteappointment(request, pk):
    request.session['deleteAppointment'] = True
    delappointmentobj = Appointment.objects.get(id=pk)
    delappointmentobj.delete()
    doctor = Doctor.objects.get(emailHash = request.session['userEmail'])
    records = doctor.doctorRecords.all()
    context = {
    "message" : "Successfully Logged In.",
    "isAuthenticated" : True,
    "user": records.order_by('-timestamp'),
    "Appointments" : Appointment.objects.all().order_by('-date')
        }
    response = response = HttpResponseRedirect(reverse('doctorappointmentsfalse'))
    return responseHeadersModifier(response)

selectedMedicineID = []
def addingMedicineData(request, selectedMedicineValue):

        try:
            SelectedMedicine = Medicine.objects.get(MedicineName = selectedMedicineValue)
            SelectedBeforeAfter = SelectedMedicine.beforeAfter
            SelectedMorning = SelectedMedicine.Morning
            SelectedAfternoon = SelectedMedicine.Afternoon
            SelectedNight = SelectedMedicine.Night
            medicineID = SelectedMedicine.pk
            selectedMedicineID.append(medicineID)
            # prescription = Prescription.objects.get(id= getPrescriptionID)
            # prescription.medicine.add(medicineID)
            data = {
                "MedBefAft" : SelectedBeforeAfter,
                "MedMorn" : SelectedMorning,
                "medAft"   : SelectedAfternoon,
                "medNight" : SelectedNight,
                # "patients" : Patient.objects.all().order_by('id'),
                # "prescPatients" : Prescription.objects.all().order_by('id'),
                # "prescMedicines" : Medicine.objects.all().order_by('MedicineName')
            }
            return JsonResponse(data)
            # response = render(request, "HealthCentre/NewPrescription.html", data)
            # return responseHeadersModifier(response)
        except Medicine.DoesNotExist:
                return JsonResponse({'error': 'Medicine not found'}, status=404)

# dummyBoolean = False
global dummyBoolean
dummyBoolean = False

def dummy(request):
    if request.method == 'GET':
        global dummyBoolean
        dummyBoolean = False
    if request.method == 'POST':
        dummyBoolean = True
    return HttpResponseRedirect(reverse("doctorprofile"))
    

def doctorprofile(request):
    #  selectedMedicineValue = ""
     if request.method == 'GET':
        request.session['writeNewPrescription'] = True

        if request.GET.get('SelectedMed') == None and request.GET.get('SelectedPat') == None:
            context = {
                    "patients" : Patient.objects.all().order_by('id'),
                    # "prescPatients" : Prescription.objects.all().order_by('id'),
                    "prescMedicines" : Medicine.objects.all().order_by('MedicineName')
                    }
            response = render(request, "HealthCentre/NewPrescription.html", context)
            return responseHeadersModifier(response)
        if request.GET.get('SelectedPat') != None and request.method == 'GET':
            PatientName = request.GET.get('SelectedPat', None)
            try: 
                selectedPatient = Patient.objects.get(name = PatientName)
                data = {
                    "patientSex" : selectedPatient.passwordHash,
                    "patientAge" : selectedPatient.rollNumber
                }
                return JsonResponse(data)
            except Patient.DoesNotExist:
                return JsonResponse({'error': 'Patient not found'}, status=404)

        if request.GET.get('SelectedMed') != None and request.method == 'GET':
    
            MedicineName = request.GET.get('SelectedMed', None)
            
            try:
                # SelectedMedicine = Medicine.objects.get(MedicineName = selectedMedicineValue)
                SelectedMedicine = Medicine.objects.get(MedicineName = MedicineName)
                medicineID = SelectedMedicine.pk
                selectedMedicineID.append(medicineID)
                SelectedBeforeAfter = SelectedMedicine.beforeAfter
                SelectedMorning = SelectedMedicine.Morning
                SelectedAfternoon = SelectedMedicine.Afternoon
                SelectedNight = SelectedMedicine.Night
                data = {
                    "MedBefAft" : SelectedBeforeAfter,
                    "MedMorn" : SelectedMorning,
                    "medAft"   : SelectedAfternoon,
                    "medNight" : SelectedNight,
                }
                return JsonResponse(data)
            except Medicine.DoesNotExist:
                return JsonResponse({'error': 'Medicine not found'}, status=404)
            
     if request.method == 'POST':
        
        if request.session['writeNewPrescription']:
            if request.POST['selectedPatient'] == "":
                prescpatient = request.POST['PatientName']
                patient = Patient.objects.create(name=prescpatient)
            else:
                prescpatient = request.POST['selectedPatient']
                patient_id = request.POST['selectedPatient'] 
                patient = Patient.objects.get(name=patient_id)

            symptoms = "dummy"#request.POST["symptoms"]
            if request.session['isLoggedIn'] and request.session['isDoctor']:
                prescdoctor = request.session['Name']
                # doctor = Doctor.objects.get(id=1)
                doctor_id = request.session['Name']
                doctor = Doctor.objects.get(name=doctor_id)
                medicine = request.POST['SelectedMedicine']
                MedicineObject = Medicine.objects.get(MedicineName = medicine)
                
                selectedMedicines = Medicine.objects.filter(id__in = selectedMedicineID)
                # MedName = MedicineObject.MedicineName
                NoOfDays = request.POST['noOfDays']
                # patient_id = request.POST['selectedPatient'] 
                patientObj = Patient.objects.get(name=prescpatient)
                prescriptiontext = "dummy"#request.POST['prescription']
                prescription = Prescription(prescribingDoctor = prescdoctor, prescribingPatient = prescpatient ,doctor = doctor, patient= patient, symptoms = symptoms, prescriptionText = prescriptiontext, NoOfDays = NoOfDays) #medicine = medicine,
                prescription.save()
                global getPrescriptionID 
                getPrescriptionID = prescription.pk
                prescription.medicine.set(selectedMedicines)
                wpnumber = patientObj.contactNumber
                global dummyBoolean
                if dummyBoolean == True:
                    sendPdfinWhatsapp(wpnumber)
            context = {
                    "prescriptions" : Prescription.objects.all().order_by('timestamp')
                }
        response = render(request, "HealthCentre/prescriptionportal.html", context)
        return responseHeadersModifier(response)

def deleteprescription(request, pk):
    # request.session['deleteAppointment'] = True
    delprescriptionobj = Prescription.objects.get(id=pk)
    delprescriptionobj.delete()
    doctor = Doctor.objects.get(emailHash = request.session['userEmail'])
    records = doctor.doctorRecords.all()
    context = {
    "message" : "Successfully Logged In.",
    "isAuthenticated" : True,
    "user": records.order_by('-timestamp'),
    "prescriptions" : Prescription.objects.all().order_by('-timestamp')
        }
    # response = render(request, 'HealthCentre/prescriptionPortal.html', context)
    response = response = HttpResponseRedirect(reverse('login'))
    return responseHeadersModifier(response)

def onlineprescription(request):
    """Function to submit online prescription request to doctor."""

    # Calling session variables checker
    request = requestSessionInitializedChecker(request)

    # If the request method is get
    if request.method == "GET":

        # If the user is logged in
        if request.session['isLoggedIn']:

            # Portal only for patient prescription request submission, not for doctors
            if request.session['isDoctor']:

                # Storing message inside context variable
                # context = {
                #         "message":"Only for patients."
                # }

                # Editing response headers so as to ignore cached versions of pages
                # response = render(request, "HealthCentre/prescriptionPortal.html", context)
                response = render(request, "HealthCentre/userDoctorProfilePortal.html")
                return responseHeadersModifier(response)

            # If the user is a patient
            else:

                # Storing available doctors inside context variable
                context = {
                    "doctors" : Doctor.objects.all().order_by('specialization')
                }

                # Editing response headers so as to ignore cached versions of pages
                response = render(request, "HealthCentre/prescriptionPortal.html", context)
                return responseHeadersModifier(response)

        # If the user is not logged in
        else:

            # Storing message inside context variable
            context = {
                    "message":"Please Login First."
            }

            # Editing response headers so as to ignore cached versions of pages
            response = render(request, "HealthCentre/prescriptionPortal.html", context)
            return responseHeadersModifier(response)

    # If the user is posting the prescription request
    elif request.method == "POST":

        # Accepting only if the user is logged in
        if request.session['isLoggedIn']:

            # If the prescription is being submitted back by a doctor
            if request.session['isDoctor']:

                # Extracting information from post request
                prescriptionText = request.POST['prescription']

                # Updating the prescription and saving it
                prescription = Prescription.objects.get(pk = request.POST['prescriptionID'])
                prescription.prescriptionText = prescriptionText
                prescription.isCompleted = True
                prescription.isNew = True
                prescription.save()

                # Getting the records of the doctor
                records = Doctor.objects.get(emailHash = request.session['userEmail']).doctorRecords.all()

                # Storing required information inside context variable
                context = {
                    "user" : records,
                    "successPrescriptionMessage" : "Prescription Successfully Submitted."
                }

                # Editing response headers so as to ignore cached versions of pages
                response = render(request, "HealthCentre/userDoctorProfilePortal.html", context)
                return responseHeadersModifier(response)

            # Else if the patient is submitting prescription request
            else:

                # Extracting information from post request and getting the corresponding doctor
                doctor = Doctor.objects.get(pk = request.POST["doctor"])
                symptoms = request.POST["symptoms"]

                # Saving the prescription under the concerned doctor
                prescription = Prescription(doctor = doctor, patient = Patient.objects.get(emailHash = request.session['userEmail']), symptoms = symptoms)
                prescription.save()

                # Storing information inside context variable
                context = {
                    "successPrescriptionMessage" : "Prescription Successfully Requested.",
                    "doctors"  : Doctor.objects.all().order_by('specialization')
                }

                # Editing response headers so as to ignore cached versions of pages
                # response = render(request, "HealthCentre/userDoctorProfilePortal.html", context)
                response = render(request, "HealthCentre/prescriptionPortal.html", context)
                return responseHeadersModifier(response)

        # Else if the user is not logged in
        else:

            # Storing information inside context variable
            context = {
                    "successPrescriptionMessage":"Please Login First.",
            }

            # Editing response headers so as to ignore cached versions of pages
            response = render(request, "HealthCentre/loginPortal.html", context)
            return responseHeadersModifier(response)

    # For any other method of access, returning a new blank online prescription page
    else:

        # Editing response headers so as to ignore cached versions of pages
        response = render(request, "HealthCentre/prescriptionPortal.html")
        return responseHeadersModifier(response)

def responseHeadersModifier(response):
    """Funtion to edit response headers so that no cached versions can be viewed. Returns the modified response."""
    response["Cache-Control"] = "no-cache, no-store, must-revalidate"
    response["Pragma"] = "no-cache"
    response["Expires"] = "0"
    return response

def requestSessionInitializedChecker(request):
    """Function to initialize request sessions if they don't exist."""

    # Try except for KeyError
    try:
        # Checking if session variables exist
        if request.session['goToAppointmentsPage'] and request.session['CreatenewAppointment'] and request.session['isDoctor'] and request.session['isLoggedIn'] and request.session['userEmail'] and request.session['Name'] and request.session['numberNewPrescriptions'] and request.session['writeNewPrescription']:
            # Do nothing if they do exist
            pass
    except:
        # Initialize request variables if they don't exist
        request.session['isDoctor'] = ""
        request.session['isLoggedIn'] = False
        request.session['userEmail'] = ""
        request.session['Name'] = ""
        request.session['numberNewPrescriptions'] = ""
        request.session['writeNewPrescription'] = False
        request.session['CreatenewAppointment'] = False
        request.session['goToAppointmentsPage'] = False
        request.session['appointmentEdit'] = False
        dummyBoolean = False
 
    # Returning request
    return request

def whatsappNotification():
    
    while True:
        currentDateTime = datetime.now()
        currentHourObj = datetime.strftime(currentDateTime, "%H")
        currentMinuteObj = datetime.strftime(currentDateTime, "%M")
        currentDateObj = datetime.strftime(currentDateTime, "%d")
        currentMonthObj = datetime.strftime(currentDateTime, "%m")
        currentYearObj = datetime.strftime(currentDateTime, "%Y")

        strCurrenTimeObj = currentHourObj + currentMinuteObj
        strCurrentDateObj = currentDateObj + currentMonthObj + currentYearObj 

        plusthreehours=currentDateTime + timedelta(hours=3)
        hourObject = datetime.strftime(plusthreehours, "%H")
        minuteObject = datetime.strftime(plusthreehours, "%M")

        strTimeObject = hourObject + minuteObject
        datetimeObject = datetime.strptime(strTimeObject, "%H%M")
        datetimeObject = datetimeObject.time()

        currentDateTimeObj = datetime.strptime(strCurrenTimeObj, "%H%M")
        currentDateTimeObj = currentDateTimeObj.time()

        currentDate = datetime.strptime(strCurrentDateObj, "%d%m%Y")
        currentDate = currentDate.date()
        # getappointmentTime = Appointment.objects.get(time=datetimeObject)
        try:
            getAllAppointmentTime = Appointment.objects.filter(time=datetimeObject)

        except Appointment.DoesNotExist:
            for getappointmentTime in getAllAppointmentTime:
                getappointmentTime = None
        # time = models.TimeField(default=timezone.now)
        # getappointmentTime = Appointment.objects.get_or_(time=datetimeObject)
        for getappointmentTime in getAllAppointmentTime:
            if not getappointmentTime == None:
                AppointmentTime = getappointmentTime.time
                AppointmentDate = getappointmentTime.date
                patientName=getappointmentTime.appointmentpatient
                patientDetail = Patient.objects.get(name=patientName)
                patientNumber = patientDetail.contactNumber
                if (AppointmentDate == currentDate):
                    whatsappApi(patientName, patientNumber, AppointmentTime, AppointmentDate)
                    time.sleep(60)
        # while True:
        #     updateExcel()
        #     time.sleep(900)
def backgroundTask():
    thread = threading.Thread(target=whatsappNotification)
    thread.daemon = True
    thread.start()
backgroundTask()

def backgroundtastForUpdatingExcel():
    xlthread = threading.Thread(target= updateExcel)
    xlthread.daemon = True
    xlthread.start()
backgroundtastForUpdatingExcel()

def searchAppointments(request):
    if request.method == "POST":

        searchQuery = request.POST["searchQuery"]

        searchFilterAppointments = Appointment.objects.filter(Q(appointmentpatient__contains = searchQuery) |
                                                            Q(appointmentdoctor__contains = searchQuery) |
                                                            Q(notes__contains = searchQuery) |
                                                            Q(date__contains = searchQuery) |
                                                            Q(time__contains = searchQuery) |
                                                            Q(subject__contains = searchQuery))
        context = {
            'searchAppointmentPatients' : searchFilterAppointments.order_by('appointmentpatient')
        }

        response = render(request, "HealthCentre/appointmentsPortal.html", context)
        return responseHeadersModifier(response)
    
def searchPrescriptions(request):
    if request.method == "POST":

        searchQuery = request.POST["searchQuery"]

        searchFilterPrescriptions = Prescription.objects.filter(Q(prescribingPatient__contains = searchQuery) | 
                                                                Q(medicine__contains = searchQuery) |
                                                                Q(timestamp__contains = searchQuery))

        context = {
            'searchPrescriptionPatients' : searchFilterPrescriptions.order_by('prescribingPatient')
        }

        response = render(request, "HealthCentre/prescriptionsPortal.html", context)
        return responseHeadersModifier(response)

def generatePDF(request):
    if request.method == "GET":
        pyautogui.hotkey('ctrl', 'p')
    # time.sleep(60)
    return HttpResponseRedirect(reverse("doctorprofile"))

def sendPdfinWhatsapp(wpnumber):

    curPath = os.getcwd() #"D:\prescPDF"
    pdfPath = os.path.join(curPath, "prescPDF")
    allFilesInPath = os.listdir(pdfPath)
    # filesInPath = [allfiles for allfiles in allFilesInPath if allfiles.lower().startswith('prescPDF')]
    # if filesInPath:
    PdfFilesInPath = [file for file in allFilesInPath if file.lower().endswith('.pdf')]
    
    if not PdfFilesInPath:
        pass
    else:
        pdfFullPaths = [os.path.join(pdfPath, pdfFile)  for pdfFile in PdfFilesInPath] 
        latestPdf = max(pdfFullPaths, key=os.path.getmtime)
    
    whatsappMedia(wpnumber, latestPdf)

# def updateGoogleSheets():
#     excelFilePath = 'D:\Dental-Software\database_tables.xlsx'
#     GsWorkbook = openpyxl.load_workbook(excelFilePath)
#     GSWorksheet = GsWorkbook.active
#     GsheetUrl = 'https://docs.google.com/spreadsheets/d/e/2PACX-1vSnNGw6Mh3W9nuNkUku96aWP04fJMyTqMwJrOQsgtAAbCkdcAcafs_SH85Ve9IluvjXdulA8HZnPTXy/pubhtml'

#     for row in GSWorksheet.iter_rows(values_only = True):
#         data = ','.join(map(str, row))
#         response = requests.get(f'{GsheetUrl}?gid=0&single=true&output=csv&exportformat=csv&range=A1', data=data)
    
#         if response.status_code == 200:
#             print(f'Successfully updated: {data}')
#         else:
#             print(f'Failed to update: {data}')